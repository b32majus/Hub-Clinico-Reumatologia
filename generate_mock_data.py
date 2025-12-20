#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para generar datos ficticios realistas para Hub Clínico
Crea 60 pacientes (30 ESPA + 30 APS) con 2-5 visitas cada uno
Total esperado: ~180-210 registros de visitas
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import random
from datetime import datetime, timedelta
import math

# ================ CONSTANTES ================

TOTAL_ESPA = 30
TOTAL_APS = 30
MIN_VISITS = 2
MAX_VISITS = 5

# Pools de datos españoles realistas
NOMBRES = [
    'Juan', 'María', 'José', 'Carmen', 'Antonio', 'Isabel', 'Francisco', 'Ana',
    'Manuel', 'Teresa', 'García', 'Rosa', 'Luis', 'Dolores', 'Miguel', 'Magdalena',
    'Javier', 'Francisca', 'Andrés', 'Esperanza', 'Jorge', 'Amparo', 'Enrique', 'Pilar',
    'Ramón', 'Consuelo', 'Pedro', 'Socorro', 'Guillermo', 'Adelaida', 'Sergio', 'Concepción',
    'Alfonso', 'Filomena', 'Fernando', 'Agripina', 'Óscar', 'Jacinta', 'Ricardo', 'Elvira',
    'Emilio', 'Matilde', 'Héctor', 'Modesta', 'Julio', 'Genoveva', 'Daniel', 'Violeta',
    'Arturo', 'Antonia', 'Valentín', 'Soledad', 'Ernesto', 'Remedios', 'Adolfo', 'Josefa'
]

APELLIDOS = [
    'García', 'Martínez', 'López', 'González', 'Fernández', 'Rodríguez', 'Sánchez',
    'Pérez', 'Gómez', 'Díaz', 'Cruz', 'Moreno', 'Gutiérrez', 'Ortiz', 'Jiménez',
    'Castillo', 'Ruiz', 'Navarro', 'Ramos', 'Herrera', 'Vázquez', 'Cano', 'Molina',
    'Romero', 'León', 'Iglesias', 'Domínguez', 'Serrano', 'Flores', 'Martín', 'Delgado',
    'Cortés', 'Campos', 'Acosta', 'Medina', 'Reyes', 'Rojas', 'Carrillo', 'Espinosa',
    'Santiago', 'Vargas', 'Cabrera', 'Montoya', 'Benítez', 'Carmona', 'Santana', 'Romeral'
]

PROFESIONALES = [
    'Dr. Silvia García',
    'Dra. Carmen López',
    'Dr. Miguel Torres',
    'Dra. Ana Martínez',
    'Dr. Javier Ruiz',
    'Dra. Isabel Sánchez'
]

# Articulaciones del homúnculo (mismo orden que en homunculus.js)
ARTICULATIONS = [
    'hombro-derecho', 'hombro-izquierdo', 'codo-derecho', 'codo-izquierdo',
    'muneca-derecha', 'muneca-izquierda', 'rodilla-derecha', 'rodilla-izquierda',
    'mcf1-derecha', 'mcf2-derecha', 'mcf3-derecha', 'mcf4-derecha', 'mcf5-derecha',
    'mcf1-izquierda', 'mcf2-izquierda', 'mcf3-izquierda', 'mcf4-izquierda', 'mcf5-izquierda',
    'ifp1-derecha', 'ifp2-derecha', 'ifp3-derecha', 'ifp4-derecha', 'ifp5-derecha',
    'ifp1-izquierda', 'ifp2-izquierda', 'ifp3-izquierda', 'ifp4-izquierda', 'ifp5-izquierda'
]

DACTILITIS = [
    'dactilitis-dedo1-mano-derecha', 'dactilitis-dedo2-mano-derecha', 'dactilitis-dedo3-mano-derecha',
    'dactilitis-dedo4-mano-derecha', 'dactilitis-dedo5-mano-derecha',
    'dactilitis-dedo1-mano-izquierda', 'dactilitis-dedo2-mano-izquierda', 'dactilitis-dedo3-mano-izquierda',
    'dactilitis-dedo4-mano-izquierda', 'dactilitis-dedo5-mano-izquierda',
    'dactilitis-dedo1-pie-derecho', 'dactilitis-dedo2-pie-derecho', 'dactilitis-dedo3-pie-derecho',
    'dactilitis-dedo4-pie-derecho', 'dactilitis-dedo5-pie-derecho',
    'dactilitis-dedo1-pie-izquierdo', 'dactilisis-dedo2-pie-izquierdo', 'dactilitis-dedo3-pie-izquierdo',
    'dactilitis-dedo4-pie-izquierdo', 'dactilitis-dedo5-pie-izquierdo'
]

# Tratamientos por patología
TRATAMIENTOS_ESPA = {
    'AINEs': ['Naproxeno 500mg/12h', 'Ibuprofeno 600mg/8h', 'Etoricoxib 120mg/día'],
    'FAMEs': ['Sulfasalazina 2g/día', 'Metotrexato 15mg/semana'],
    'Biológicos': ['Adalimumab 40mg/2sem', 'Etanercept 50mg/semana', 'Secukinumab 150mg/4sem',
                   'Ixekizumab 80mg/4sem', 'Certolizumab 400mg/4sem']
}

TRATAMIENTOS_APS = {
    'AINEs': ['Naproxeno 500mg/12h', 'Ibuprofeno 600mg/8h', 'Etoricoxib 120mg/día'],
    'FAMEs': ['Metotrexato 15mg/semana', 'Leflunomida 20mg/día'],
    'Biológicos': ['Adalimumab 40mg/2sem', 'Etanercept 50mg/semana', 'Secukinumab 150mg/4sem',
                   'Ustekinumab 90mg/12sem', 'Guselkumab 100mg/4sem']
}

# ================ FUNCIONES GENERADORAS ================

def generar_fecha_nacimiento():
    """Genera fecha de nacimiento realista (30-75 años en 2025)"""
    year = random.randint(1950, 1995)
    month = random.randint(1, 12)
    day = random.randint(1, 28)
    return f"{year:04d}-{month:02d}-{day:02d}"

def generar_nombre_completo():
    """Genera nombre y apellido españoles realistas"""
    nombre = random.choice(NOMBRES)
    apellido1 = random.choice(APELLIDOS)
    apellido2 = random.choice(APELLIDOS)
    return f"{nombre} {apellido1} {apellido2}"

def generar_id_paciente(pathology, index):
    """Genera ID de paciente con mix 50%-50% 2024-2025"""
    year = random.choice([2024, 2025])
    path_code = 'ESP' if pathology == 'espa' else 'APS'
    return f"{path_code}-{year}-{str(index+1).zfill(3)}"

def generar_articulaciones(count_nad_max=12):
    """Genera array de articulaciones dolorosas realista (NAD)"""
    count_nad = random.randint(0, count_nad_max)
    nad = random.sample(ARTICULATIONS, count_nad) if count_nad > 0 else []

    # NAT es subconjunto de NAD (0-66% de NAD)
    if len(nad) > 0:
        count_nat = random.randint(0, int(len(nad) * 0.66))
        nat = random.sample(nad, count_nat) if count_nat > 0 else []
    else:
        nat = []

    return nad, nat

def generar_dactilitis():
    """Genera dactilitis realista (0-3 dedos, más manos que pies)"""
    count = random.randint(0, 3)
    if count == 0:
        return []

    mano_count = random.randint(count // 2, count)
    pie_count = count - mano_count

    dactilitis = []
    if mano_count > 0:
        dactilitis.extend(random.sample(DACTILITIS[:10], mano_count))
    if pie_count > 0:
        dactilitis.extend(random.sample(DACTILITIS[10:], pie_count))

    return dactilitis

def generar_basdai_con_distribucion_normal(mejora=False):
    """Genera BASDAI con distribución normal (media 3.5, sd 1.5)"""
    basdai = random.gauss(3.5, 1.5)
    basdai = max(1.0, min(10.0, basdai))  # Clamp entre 1 y 10

    # En seguimientos hay tendencia a mejorar 20-40%
    if mejora:
        basdai *= random.uniform(0.6, 0.8)

    return round(basdai, 1)

def generar_haq_con_distribucion_normal(mejora=False):
    """Genera HAQ con distribución normal (media 1.2, sd 0.7)"""
    haq = random.gauss(1.2, 0.7)
    haq = max(0.0, min(3.0, haq))

    if mejora:
        haq *= random.uniform(0.7, 0.85)

    return round(haq, 2)

def generar_pasi():
    """Genera PASI realista para psoriasis (2-20)"""
    return random.randint(2, 20)

def expandir_articulaciones_a_columnas(array):
    """Convierte array de articulaciones a SI/NO para cada columna"""
    return ['SI' if art in array else 'NO' for art in ARTICULATIONS]

def expandir_dactilitis_a_columnas(array):
    """Convierte array de dactilitis a SI/NO para cada columna"""
    return ['SI' if dact in array else 'NO' for dact in DACTILITIS]

# ================ GENERACIÓN DE VISITAS ================

def generar_primera_visita_espa(paciente_id, nombre, sexo, fecha_visita, datos_paciente):
    """Genera fila completa de primera visita ESPA (220 columnas)"""

    nad, nat = generar_articulaciones()
    dactilitis = generar_dactilitis()

    basdai = generar_basdai_con_distribucion_normal()
    asdas = basdai * 0.8 + random.uniform(-0.5, 0.5)  # ASDAS correlacionado con BASDAI

    fila = [
        # Identificación (5)
        paciente_id, nombre, sexo, fecha_visita.strftime('%Y-%m-%d'), 'Primera Visita',
        # Profesional y diagnóstico (6)
        datos_paciente['profesional'], 'ESPA', '',
        datos_paciente['hla_b27'], '', '',
        # Anamnesis inicial (7)
        '2024-01' if random.random() < 0.5 else '2023-06',  # Inicio_Sintomas (mes-año)
        '', 'SI' if random.random() < 0.4 else 'NO', '30', '45', 'SI' if random.random() < 0.6 else 'NO',
        'SI' if random.random() < 0.7 else 'NO',
        # NAD expandidas (30)
        *expandir_articulaciones_a_columnas(nad),
        # NAT expandidas (30)
        *expandir_articulaciones_a_columnas(nat),
        # Dactilitis expandidas (20)
        *expandir_dactilitis_a_columnas(dactilitis),
        # Totales (3)
        len(nad), len(nat), len(dactilitis),
        # Antropometría (4)
        random.randint(60, 95), random.randint(160, 185),
        round(70 / (1.75**2), 1), f"{random.randint(110, 140)}/{random.randint(70, 90)}",
        # PROs - vacío en primera visita (5)
        '', '', '', '', '',
        # Psoriasis - vacío en ESPA (5)
        '', '', '', '', '',
        # ExtraArticular (3)
        'SI' if random.random() < 0.1 else 'NO', 'SI' if random.random() < 0.15 else 'NO', 'NO',
        # Comorbilidades (8)
        'SI' if random.random() < 0.35 else 'NO', 'SI' if random.random() < 0.2 else 'NO',
        'SI' if random.random() < 0.4 else 'NO', 'SI' if random.random() < 0.05 else 'NO',
        'SI' if random.random() < 0.1 else 'NO', 'SI' if random.random() < 0.3 else 'NO',
        'SI' if random.random() < 0.15 else 'NO', 'SI' if random.random() < 0.05 else 'NO',
        # AF (4)
        '', '', '', '',
        # Tóxicos (6)
        'SI' if random.random() < 0.3 else 'NO', '', 'SI' if random.random() < 0.2 else 'NO', '',
        'SI' if random.random() < 0.1 else 'NO', '',
        # Entesitis (11)
        'SI' if random.random() < 0.2 else 'NO', 'SI' if random.random() < 0.15 else 'NO',
        'SI' if random.random() < 0.1 else 'NO', 'SI' if random.random() < 0.08 else 'NO',
        'SI' if random.random() < 0.05 else 'NO', 'SI' if random.random() < 0.2 else 'NO',
        'SI' if random.random() < 0.15 else 'NO', 'SI' if random.random() < 0.1 else 'NO',
        'SI' if random.random() < 0.08 else 'NO', 'SI' if random.random() < 0.05 else 'NO',
        'SI' if random.random() < 0.1 else 'NO',
        # Pruebas complementarias - vacío en primera (5)
        '', '', '', '', '',
        # BASDAI (7) - vacío en primera
        '', '', '', '', '', '', '',
        # ASDAS (5) - vacío en primera
        '', '', '', '', '',
        # Metrología (6) - vacío en primera
        '', '', '', '', '', '',
        # PASI (3) - vacío en ESPA
        '', '', '',
        # HAQ (9) - vacío en primera
        '', '', '', '', '', '', '', '', '',
        # LEI (7) - vacío en ESPA
        '', '', '', '', '', '', '',
        # MDA (8) - vacío en ESPA
        '', '', '', '', '', '', '', '',
        # RAPID3 (4) - vacío en ESPA
        '', '', '', '',
        # Tratamiento (3)
        datos_paciente['tratamiento_inicial'], datos_paciente['fecha_inicio'], '',
        # Continuar - vacío en primera (2)
        '', '',
        # Cambio - vacío en primera (12)
        '', '', '', '', '', '', '', '', '', '', '', '',
        # Decision_Terapeutica_SEG - vacío (1)
        '',
        # Tratamientos iniciales (6)
        datos_paciente['trat_sistemico'], datos_paciente['trat_sistemico_dosis'],
        datos_paciente['trat_fame'], datos_paciente['trat_fame_dosis'],
        datos_paciente['trat_biologico'], datos_paciente['trat_biologico_dosis'],
        # Fecha próxima revisión y comentarios (2)
        (fecha_visita + timedelta(days=180)).strftime('%Y-%m-%d'), ''
    ]

    return fila

def generar_seguimiento_espa(paciente_id, nombre, sexo, fecha_visita, datos_paciente, datos_previos):
    """Genera fila completa de seguimiento ESPA (220 columnas)"""

    nad, nat = generar_articulaciones()
    dactilitis = generar_dactilitis()

    basdai = generar_basdai_con_distribucion_normal(mejora=True)
    asdas = basdai * 0.8 + random.uniform(-0.5, 0.5)

    pcr = round(random.gauss(5, 3), 1)
    pcr = max(0, pcr)
    vsg = round(random.gauss(12, 8), 0)
    vsg = max(1, vsg)

    fila = [
        # Identificación (5)
        paciente_id, nombre, sexo, fecha_visita.strftime('%Y-%m-%d'), 'Seguimiento',
        # Profesional y diagnóstico (6)
        datos_paciente['profesional'], 'ESPA', '',
        datos_paciente['hla_b27'], '', '',
        # Anamnesis - vacío en seguimiento (7)
        '', '', '', '', '', '', '',
        # NAD expandidas (30)
        *expandir_articulaciones_a_columnas(nad),
        # NAT expandidas (30)
        *expandir_articulaciones_a_columnas(nat),
        # Dactilitis expandidas (20)
        *expandir_dactilitis_a_columnas(dactilitis),
        # Totales (3)
        len(nad), len(nat), len(dactilitis),
        # Antropometría (4)
        random.randint(60, 95), random.randint(160, 185),
        round(70 / (1.75**2), 1), f"{random.randint(110, 140)}/{random.randint(70, 90)}",
        # PROs (5)
        round(basdai, 1), round(basdai * 0.7, 1), round(basdai * 0.8, 1),
        random.randint(10, 60), round(random.uniform(2, 8), 1),
        # Psoriasis - vacío en ESPA (5)
        '', '', '', '', '',
        # ExtraArticular (3)
        'SI' if random.random() < 0.1 else 'NO', 'SI' if random.random() < 0.15 else 'NO', 'NO',
        # Comorbilidades (8)
        datos_paciente['comorbilidades']['HTA'],
        datos_paciente['comorbilidades']['DM'],
        datos_paciente['comorbilidades']['DLP'],
        'NO',  # ECV
        'SI' if random.random() < 0.1 else 'NO',  # Gastritis
        'SI' if random.random() < 0.3 else 'NO',  # Obesidad
        'SI' if random.random() < 0.15 else 'NO',  # Osteoporosis
        'SI' if random.random() < 0.05 else 'NO',  # Gota
        # AF - vacío en seguimiento (4)
        '', '', '', '',
        # Tóxicos - vacío en seguimiento (6)
        '', '', '', '', '', '',
        # Entesitis - vacío en seguimiento (11)
        '', '', '', '', '', '', '', '', '', '', '',
        # Pruebas complementarias (5)
        round(pcr, 1), int(vsg), '', '', '',
        # BASDAI (7)
        random.randint(1, 7), random.randint(1, 7), random.randint(1, 7),
        random.randint(1, 7), random.randint(1, 7), random.randint(1, 7),
        round(basdai, 1),
        # ASDAS (5)
        'dolor_espalda_' + str(random.randint(1, 10)),
        random.randint(10, 180),
        round(basdai, 1),
        round(pcr, 1),
        round(vsg, 1),
        # Metrología (6)
        round(random.uniform(3, 7), 1), random.randint(40, 80),
        round(random.uniform(5, 25), 1), round(random.uniform(40, 80), 1),
        round(random.uniform(2, 5), 1), random.randint(30, 50),
        # PASI - vacío en ESPA (3)
        '', '', '',
        # HAQ - vacío en ESPA (9)
        '', '', '', '', '', '', '', '', '',
        # LEI - vacío en ESPA (7)
        '', '', '', '', '', '', '',
        # MDA - vacío en ESPA (8)
        '', '', '', '', '', '', '', '',
        # RAPID3 - vacío en ESPA (4)
        '', '', '', '',
        # Tratamiento (3)
        datos_paciente['tratamiento_actual'], datos_paciente['fecha_inicio'], '',
        # Continuar (2)
        'SI' if random.random() < 0.6 else 'NO', 'SI' if random.random() < 0.3 else 'NO',
        # Cambio (12)
        '', '', '', '', '', '', '', '', '', '', '', '',
        # Decision_Terapeutica_SEG (1)
        'CONTINUAR' if random.random() < 0.6 else random.choice(['AUMENTAR_DOSIS', 'CAMBIAR']),
        # Tratamientos iniciales - vacío en seguimiento (6)
        '', '', '', '', '', '',
        # Fecha próxima revisión y comentarios (2)
        (fecha_visita + timedelta(days=180)).strftime('%Y-%m-%d'), ''
    ]

    return fila

# ================ GENERACIÓN DE VISITAS APS ================

def generar_primera_visita_aps(paciente_id, nombre, sexo, fecha_visita, datos_paciente):
    """Genera fila completa de primera visita APS (220 columnas)"""

    nad, nat = generar_articulaciones(count_nad_max=10)  # Menos articulaciones en APS
    dactilitis = generar_dactilitis()

    haq = generar_haq_con_distribucion_normal()
    pasi = generar_pasi()

    fila = [
        # Identificación (5)
        paciente_id, nombre, sexo, fecha_visita.strftime('%Y-%m-%d'), 'Primera Visita',
        # Profesional y diagnóstico (6)
        datos_paciente['profesional'], 'APS', '',
        'NO' if random.random() < 0.7 else 'SI',
        'SI' if random.random() < 0.3 else 'NO',
        'SI' if random.random() < 0.4 else 'NO',
        # Anamnesis (7)
        '2024-01' if random.random() < 0.5 else '2023-06',
        '2023-06' if random.random() < 0.7 else '2022-12',  # Inicio_Psoriasis
        'SI' if random.random() < 0.3 else 'NO',
        '30', '45', 'SI' if random.random() < 0.4 else 'NO',
        'SI' if random.random() < 0.5 else 'NO',
        # NAD expandidas (30)
        *expandir_articulaciones_a_columnas(nad),
        # NAT expandidas (30)
        *expandir_articulaciones_a_columnas(nat),
        # Dactilitis expandidas (20)
        *expandir_dactilitis_a_columnas(dactilitis),
        # Totales (3)
        len(nad), len(nat), len(dactilitis),
        # Antropometría (4)
        random.randint(55, 95), random.randint(155, 185),
        round(70 / (1.75**2), 1), f"{random.randint(110, 140)}/{random.randint(70, 90)}",
        # PROs - vacío en primera (5)
        '', '', '', '', '',
        # Psoriasis (5)
        'SI' if random.random() < 0.7 else 'NO',
        'SI' if random.random() < 0.4 else 'NO',
        'SI' if random.random() < 0.6 else 'NO',
        'SI' if random.random() < 0.3 else 'NO',
        'SI' if random.random() < 0.2 else 'NO',
        # ExtraArticular (3)
        'SI' if random.random() < 0.1 else 'NO',
        'SI' if random.random() < 0.05 else 'NO',
        'SI' if random.random() < 0.2 else 'NO',
        # Comorbilidades (8)
        'SI' if random.random() < 0.35 else 'NO', 'SI' if random.random() < 0.2 else 'NO',
        'SI' if random.random() < 0.4 else 'NO', 'SI' if random.random() < 0.05 else 'NO',
        'SI' if random.random() < 0.1 else 'NO', 'SI' if random.random() < 0.3 else 'NO',
        'SI' if random.random() < 0.15 else 'NO', 'SI' if random.random() < 0.05 else 'NO',
        # AF (4)
        '', '', '', '',
        # Tóxicos (6)
        'SI' if random.random() < 0.3 else 'NO', '', 'SI' if random.random() < 0.2 else 'NO', '',
        'SI' if random.random() < 0.1 else 'NO', '',
        # Entesitis (11)
        'SI' if random.random() < 0.2 else 'NO', 'SI' if random.random() < 0.15 else 'NO',
        'SI' if random.random() < 0.15 else 'NO', 'SI' if random.random() < 0.1 else 'NO',
        'SI' if random.random() < 0.08 else 'NO', 'SI' if random.random() < 0.2 else 'NO',
        'SI' if random.random() < 0.15 else 'NO', 'SI' if random.random() < 0.15 else 'NO',
        'SI' if random.random() < 0.1 else 'NO', 'SI' if random.random() < 0.08 else 'NO',
        'SI' if random.random() < 0.1 else 'NO',
        # Pruebas - vacío en primera (5)
        '', '', '', '', '',
        # BASDAI - vacío en APS (7)
        '', '', '', '', '', '', '',
        # ASDAS - vacío en primera (5)
        '', '', '', '', '',
        # Metrología - vacío en APS (6)
        '', '', '', '', '', '',
        # PASI (3)
        pasi, round(len(nad) * 2, 1), f"Afectación {'leve' if pasi < 10 else 'moderada' if pasi < 15 else 'grave'}",
        # HAQ - vacío en primera (9)
        '', '', '', '', '', '', '', '', '',
        # LEI - vacío en primera (7)
        '', '', '', '', '', '', '',
        # MDA - vacío en primera (8)
        '', '', '', '', '', '', '', '',
        # RAPID3 - vacío en primera (4)
        '', '', '', '',
        # Tratamiento (3)
        datos_paciente['tratamiento_inicial'], datos_paciente['fecha_inicio'], '',
        # Continuar - vacío (2)
        '', '',
        # Cambio - vacío (12)
        '', '', '', '', '', '', '', '', '', '', '', '',
        # Decision_Terapeutica_SEG - vacío (1)
        '',
        # Tratamientos iniciales (6)
        datos_paciente['trat_sistemico'], datos_paciente['trat_sistemico_dosis'],
        datos_paciente['trat_fame'], datos_paciente['trat_fame_dosis'],
        datos_paciente['trat_biologico'], datos_paciente['trat_biologico_dosis'],
        # Fecha próxima revisión y comentarios (2)
        (fecha_visita + timedelta(days=180)).strftime('%Y-%m-%d'), ''
    ]

    return fila

def generar_seguimiento_aps(paciente_id, nombre, sexo, fecha_visita, datos_paciente, datos_previos):
    """Genera fila completa de seguimiento APS (220 columnas)"""

    nad, nat = generar_articulaciones(count_nad_max=10)
    dactilitis = generar_dactilitis()

    haq = generar_haq_con_distribucion_normal(mejora=True)
    pasi = max(1, generar_pasi() - random.randint(0, 5))

    pcr = round(random.gauss(5, 3), 1)
    pcr = max(0, pcr)
    vsg = round(random.gauss(12, 8), 0)
    vsg = max(1, vsg)

    # LEI (7 columnas) - solo APS
    lei_campos = [
        round(random.uniform(0, 2), 0),  # LEI_Epicondilo_Lat_Izq
        round(random.uniform(0, 2), 0),  # LEI_Epicondilo_Lat_Der
        round(random.uniform(0, 2), 0),  # LEI_Epicondilo_Med_Izq
        round(random.uniform(0, 2), 0),  # LEI_Epicondilo_Med_Der
        round(random.uniform(0, 2), 0),  # LEI_Aquiles_Izq
        round(random.uniform(0, 2), 0),  # LEI_Aquiles_Der
        sum([random.randint(0, 2) for _ in range(7)]) // 7  # LEI_Score
    ]

    # MDA (8 columnas) - solo APS
    mda_cumple = len(nad) <= 4 and haq <= 1.5 and len(dactilitis) == 0 and pasi <= 10

    fila = [
        # Identificación (5)
        paciente_id, nombre, sexo, fecha_visita.strftime('%Y-%m-%d'), 'Seguimiento',
        # Profesional y diagnóstico (6)
        datos_paciente['profesional'], 'APS', '',
        'NO' if random.random() < 0.7 else 'SI',
        'SI' if random.random() < 0.3 else 'NO',
        'SI' if random.random() < 0.4 else 'NO',
        # Anamnesis - vacío (7)
        '', '', '', '', '', '', '',
        # NAD expandidas (30)
        *expandir_articulaciones_a_columnas(nad),
        # NAT expandidas (30)
        *expandir_articulaciones_a_columnas(nat),
        # Dactilitis expandidas (20)
        *expandir_dactilitis_a_columnas(dactilitis),
        # Totales (3)
        len(nad), len(nat), len(dactilitis),
        # Antropometría (4)
        random.randint(55, 95), random.randint(155, 185),
        round(70 / (1.75**2), 1), f"{random.randint(110, 140)}/{random.randint(70, 90)}",
        # PROs (5)
        round(haq * 1.2, 1), round(haq, 1), round(haq * 0.8, 1),
        random.randint(10, 60), round(random.uniform(2, 8), 1),
        # Psoriasis - vacío en seguimiento (5)
        '', '', '', '', '',
        # ExtraArticular (3)
        'SI' if random.random() < 0.1 else 'NO',
        'SI' if random.random() < 0.05 else 'NO',
        'SI' if random.random() < 0.2 else 'NO',
        # Comorbilidades (8)
        datos_paciente['comorbilidades']['HTA'],
        datos_paciente['comorbilidades']['DM'],
        datos_paciente['comorbilidades']['DLP'],
        'NO',
        'SI' if random.random() < 0.1 else 'NO',
        'SI' if random.random() < 0.3 else 'NO',
        'SI' if random.random() < 0.15 else 'NO',
        'SI' if random.random() < 0.05 else 'NO',
        # AF - vacío (4)
        '', '', '', '',
        # Tóxicos - vacío (6)
        '', '', '', '', '', '',
        # Entesitis - vacío (11)
        '', '', '', '', '', '', '', '', '', '', '',
        # Pruebas (5)
        round(pcr, 1), int(vsg), '', '', '',
        # BASDAI - vacío en APS (7)
        '', '', '', '', '', '', '',
        # ASDAS - vacío en seguimiento (5)
        '', '', '', '', '',
        # Metrología - vacío en APS (6)
        '', '', '', '', '', '',
        # PASI (3)
        pasi, round(len(nad) * 1.5, 1), f"Afectación {'leve' if pasi < 10 else 'moderada' if pasi < 15 else 'grave'}",
        # HAQ (9)
        round(haq, 2), round(haq * 0.9, 2), round(haq * 1.1, 2), round(haq * 0.8, 2),
        round(haq * 0.95, 2), round(haq * 1.05, 2), round(haq * 0.85, 2),
        round(haq * 1.15, 2), round(haq, 2),
        # LEI (7)
        *lei_campos,
        # MDA (8)
        len(nad), len(nat), pasi, round(haq, 1), 'SI' if mda_cumple else 'NO',
        len(dactilitis), 'SI' if mda_cumple else 'NO', 'SI' if mda_cumple else 'NO',
        # RAPID3 (4)
        round(haq * 10, 1), round(haq * 8, 1), round(haq * 9, 1), round(haq * 25, 1),
        # Tratamiento (3)
        datos_paciente['tratamiento_actual'], datos_paciente['fecha_inicio'], '',
        # Continuar (2)
        'SI' if random.random() < 0.6 else 'NO',
        'SI' if random.random() < 0.3 else 'NO',
        # Cambio (12)
        '', '', '', '', '', '', '', '', '', '', '', '',
        # Decision_Terapeutica_SEG (1)
        'CONTINUAR' if random.random() < 0.6 else random.choice(['AUMENTAR_DOSIS', 'CAMBIAR']),
        # Tratamientos iniciales - vacío (6)
        '', '', '', '', '', '',
        # Fecha próxima revisión y comentarios (2)
        (fecha_visita + timedelta(days=180)).strftime('%Y-%m-%d'), ''
    ]

    return fila

# ================ FUNCIÓN PRINCIPAL ================

def generar_base_datos():
    """Genera Hub_Clinico_Maestro.xlsx con 60 pacientes ficticios"""

    print("Cargando Hub_Clinico_Maestro.xlsx...")
    wb = openpyxl.load_workbook('Hub_Clinico_Maestro.xlsx')
    ws_espa = wb['ESPA']
    ws_aps = wb['APS']

    row_espa = 2
    row_aps = 2

    total_visitas_espa = 0
    total_visitas_aps = 0

    # ========== GENERAR 30 PACIENTES ESPA ==========
    print("Generando 30 pacientes ESPA...")

    for i in range(TOTAL_ESPA):
        paciente_id = generar_id_paciente('espa', i)
        nombre = generar_nombre_completo()
        sexo = random.choice(['Hombre', 'Mujer'])
        num_visitas = random.randint(MIN_VISITS, MAX_VISITS)
        fecha_primera = datetime(2024, 1, 1) + timedelta(days=random.randint(0, 365))

        # Datos constantes del paciente
        datos_paciente = {
            'profesional': random.choice(PROFESIONALES),
            'hla_b27': 'Positivo' if random.random() < 0.8 else 'Negativo',
            'comorbilidades': {
                'HTA': 'SI' if random.random() < 0.35 else 'NO',
                'DM': 'SI' if random.random() < 0.2 else 'NO',
                'DLP': 'SI' if random.random() < 0.4 else 'NO'
            }
        }

        # Escalada terapéutica
        escalada = random.random()
        if escalada < 0.4:
            datos_paciente['tratamiento_inicial'] = random.choice(TRATAMIENTOS_ESPA['AINEs'])
            datos_paciente['trat_sistemico'] = random.choice(TRATAMIENTOS_ESPA['AINEs'])
            datos_paciente['trat_sistemico_dosis'] = '500mg/12h'
            datos_paciente['trat_fame'] = ''
            datos_paciente['trat_fame_dosis'] = ''
            datos_paciente['trat_biologico'] = ''
            datos_paciente['trat_biologico_dosis'] = ''
        elif escalada < 0.7:
            aine = random.choice(TRATAMIENTOS_ESPA['AINEs'])
            fame = random.choice(TRATAMIENTOS_ESPA['FAMEs'])
            datos_paciente['tratamiento_inicial'] = f"{aine} + {fame}"
            datos_paciente['trat_sistemico'] = aine
            datos_paciente['trat_sistemico_dosis'] = '500mg/12h'
            datos_paciente['trat_fame'] = fame
            datos_paciente['trat_fame_dosis'] = '2g/día' if 'Sulfasalazina' in fame else '15mg/sem'
            datos_paciente['trat_biologico'] = ''
            datos_paciente['trat_biologico_dosis'] = ''
        else:
            aine = random.choice(TRATAMIENTOS_ESPA['AINEs'])
            bio = random.choice(TRATAMIENTOS_ESPA['Biológicos'])
            datos_paciente['tratamiento_inicial'] = f"{aine} + {bio}"
            datos_paciente['trat_sistemico'] = aine
            datos_paciente['trat_sistemico_dosis'] = '500mg/12h'
            datos_paciente['trat_fame'] = ''
            datos_paciente['trat_fame_dosis'] = ''
            datos_paciente['trat_biologico'] = bio.split()[0]
            datos_paciente['trat_biologico_dosis'] = ' '.join(bio.split()[1:])

        datos_paciente['fecha_inicio'] = fecha_primera.strftime('%Y-%m-%d')

        # Primera visita
        fila = generar_primera_visita_espa(paciente_id, nombre, sexo, fecha_primera, datos_paciente)
        for col, valor in enumerate(fila, 1):
            ws_espa.cell(row=row_espa, column=col, value=valor)
        row_espa += 1
        total_visitas_espa += 1

        # Cambio de tratamiento en algunos pacientes para seguimientos
        cambio_prob = 0.4

        # Visitas de seguimiento
        for v in range(1, num_visitas):
            fecha_visita = fecha_primera + timedelta(days=random.randint(90 + v*90, 180 + v*90))

            # Algunos pacientes cambian de tratamiento en seguimientos
            if random.random() < cambio_prob:
                nuevo_trat = random.choice(TRATAMIENTOS_ESPA['Biológicos'])
                datos_paciente['tratamiento_actual'] = nuevo_trat
            else:
                datos_paciente['tratamiento_actual'] = datos_paciente['tratamiento_inicial']

            fila = generar_seguimiento_espa(paciente_id, nombre, sexo, fecha_visita, datos_paciente, None)
            for col, valor in enumerate(fila, 1):
                ws_espa.cell(row=row_espa, column=col, value=valor)
            row_espa += 1
            total_visitas_espa += 1

    # ========== GENERAR 30 PACIENTES APS ==========
    print("Generando 30 pacientes APS...")

    for i in range(TOTAL_APS):
        paciente_id = generar_id_paciente('aps', i)
        nombre = generar_nombre_completo()
        sexo = random.choice(['Hombre', 'Mujer'])
        num_visitas = random.randint(MIN_VISITS, MAX_VISITS)
        fecha_primera = datetime(2024, 1, 1) + timedelta(days=random.randint(0, 365))

        datos_paciente = {
            'profesional': random.choice(PROFESIONALES),
            'hla_b27': 'Negativo' if random.random() < 0.7 else 'Positivo',
            'comorbilidades': {
                'HTA': 'SI' if random.random() < 0.35 else 'NO',
                'DM': 'SI' if random.random() < 0.2 else 'NO',
                'DLP': 'SI' if random.random() < 0.4 else 'NO'
            }
        }

        # Escalada terapéutica (similar a ESPA)
        escalada = random.random()
        if escalada < 0.4:
            datos_paciente['tratamiento_inicial'] = random.choice(TRATAMIENTOS_APS['AINEs'])
            datos_paciente['trat_sistemico'] = random.choice(TRATAMIENTOS_APS['AINEs'])
            datos_paciente['trat_sistemico_dosis'] = '500mg/12h'
            datos_paciente['trat_fame'] = ''
            datos_paciente['trat_fame_dosis'] = ''
            datos_paciente['trat_biologico'] = ''
            datos_paciente['trat_biologico_dosis'] = ''
        elif escalada < 0.7:
            aine = random.choice(TRATAMIENTOS_APS['AINEs'])
            fame = random.choice(TRATAMIENTOS_APS['FAMEs'])
            datos_paciente['tratamiento_inicial'] = f"{aine} + {fame}"
            datos_paciente['trat_sistemico'] = aine
            datos_paciente['trat_sistemico_dosis'] = '500mg/12h'
            datos_paciente['trat_fame'] = fame
            datos_paciente['trat_fame_dosis'] = '20mg/día' if 'Leflunomida' in fame else '15mg/sem'
            datos_paciente['trat_biologico'] = ''
            datos_paciente['trat_biologico_dosis'] = ''
        else:
            aine = random.choice(TRATAMIENTOS_APS['AINEs'])
            bio = random.choice(TRATAMIENTOS_APS['Biológicos'])
            datos_paciente['tratamiento_inicial'] = f"{aine} + {bio}"
            datos_paciente['trat_sistemico'] = aine
            datos_paciente['trat_sistemico_dosis'] = '500mg/12h'
            datos_paciente['trat_fame'] = ''
            datos_paciente['trat_fame_dosis'] = ''
            datos_paciente['trat_biologico'] = bio.split()[0]
            datos_paciente['trat_biologico_dosis'] = ' '.join(bio.split()[1:])

        datos_paciente['fecha_inicio'] = fecha_primera.strftime('%Y-%m-%d')

        # Primera visita
        fila = generar_primera_visita_aps(paciente_id, nombre, sexo, fecha_primera, datos_paciente)
        for col, valor in enumerate(fila, 1):
            ws_aps.cell(row=row_aps, column=col, value=valor)
        row_aps += 1
        total_visitas_aps += 1

        # Visitas de seguimiento
        for v in range(1, num_visitas):
            fecha_visita = fecha_primera + timedelta(days=random.randint(90 + v*90, 180 + v*90))

            if random.random() < cambio_prob:
                nuevo_trat = random.choice(TRATAMIENTOS_APS['Biológicos'])
                datos_paciente['tratamiento_actual'] = nuevo_trat
            else:
                datos_paciente['tratamiento_actual'] = datos_paciente['tratamiento_inicial']

            fila = generar_seguimiento_aps(paciente_id, nombre, sexo, fecha_visita, datos_paciente, None)
            for col, valor in enumerate(fila, 1):
                ws_aps.cell(row=row_aps, column=col, value=valor)
            row_aps += 1
            total_visitas_aps += 1

    # Guardar archivo
    print(f"Guardando Hub_Clinico_Maestro.xlsx...")
    wb.save('Hub_Clinico_Maestro.xlsx')

    print("\n" + "="*60)
    print("✅ ÉXITO: Base de datos generada correctamente")
    print("="*60)
    print(f"ESPA: {TOTAL_ESPA} pacientes, {total_visitas_espa} visitas")
    print(f"APS:  {TOTAL_APS} pacientes, {total_visitas_aps} visitas")
    print(f"TOTAL: {TOTAL_ESPA + TOTAL_APS} pacientes, {total_visitas_espa + total_visitas_aps} visitas")
    print("="*60)

if __name__ == '__main__':
    generar_base_datos()
