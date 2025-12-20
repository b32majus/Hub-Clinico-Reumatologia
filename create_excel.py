import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Crear un nuevo workbook
wb = openpyxl.Workbook()

# Eliminar la hoja predeterminada
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Definir los headers
headers = [
    'ID_Paciente', 'Nombre_Paciente', 'Sexo', 'Fecha_Visita', 'Tipo_Visita',
    'Profesional', 'Diagnostico_Primario', 'Diagnostico_Secundario', 'HLA_B27', 'FR', 'APCC',
    'Inicio_Sintomas', 'Inicio_Psoriasis', 'Dolor_Axial', 'Rigidez_Matutina', 'Duracion_Rigidez', 'Irradiacion_Nalgas', 'Clinica_Axial_Presente',
    # NAD individuales (30)
    'NAD_hombro_derecho', 'NAD_hombro_izquierdo', 'NAD_codo_derecho', 'NAD_codo_izquierdo',
    'NAD_muneca_derecha', 'NAD_muneca_izquierda', 'NAD_rodilla_derecha', 'NAD_rodilla_izquierda',
    'NAD_mcf1_derecha', 'NAD_mcf2_derecha', 'NAD_mcf3_derecha', 'NAD_mcf4_derecha', 'NAD_mcf5_derecha',
    'NAD_mcf1_izquierda', 'NAD_mcf2_izquierda', 'NAD_mcf3_izquierda', 'NAD_mcf4_izquierda', 'NAD_mcf5_izquierda',
    'NAD_ifp1_derecha', 'NAD_ifp2_derecha', 'NAD_ifp3_derecha', 'NAD_ifp4_derecha', 'NAD_ifp5_derecha',
    'NAD_ifp1_izquierda', 'NAD_ifp2_izquierda', 'NAD_ifp3_izquierda', 'NAD_ifp4_izquierda', 'NAD_ifp5_izquierda',
    # NAT individuales (30)
    'NAT_hombro_derecho', 'NAT_hombro_izquierdo', 'NAT_codo_derecho', 'NAT_codo_izquierdo',
    'NAT_muneca_derecha', 'NAT_muneca_izquierda', 'NAT_rodilla_derecha', 'NAT_rodilla_izquierda',
    'NAT_mcf1_derecha', 'NAT_mcf2_derecha', 'NAT_mcf3_derecha', 'NAT_mcf4_derecha', 'NAT_mcf5_derecha',
    'NAT_mcf1_izquierda', 'NAT_mcf2_izquierda', 'NAT_mcf3_izquierda', 'NAT_mcf4_izquierda', 'NAT_mcf5_izquierda',
    'NAT_ifp1_derecha', 'NAT_ifp2_derecha', 'NAT_ifp3_derecha', 'NAT_ifp4_derecha', 'NAT_ifp5_derecha',
    'NAT_ifp1_izquierda', 'NAT_ifp2_izquierda', 'NAT_ifp3_izquierda', 'NAT_ifp4_izquierda', 'NAT_ifp5_izquierda',
    # Dactilitis individuales (20)
    'DACT_dedo1_mano_derecha', 'DACT_dedo2_mano_derecha', 'DACT_dedo3_mano_derecha', 'DACT_dedo4_mano_derecha', 'DACT_dedo5_mano_derecha',
    'DACT_dedo1_mano_izquierda', 'DACT_dedo2_mano_izquierda', 'DACT_dedo3_mano_izquierda', 'DACT_dedo4_mano_izquierda', 'DACT_dedo5_mano_izquierda',
    'DACT_dedo1_pie_derecho', 'DACT_dedo2_pie_derecho', 'DACT_dedo3_pie_derecho', 'DACT_dedo4_pie_derecho', 'DACT_dedo5_pie_derecho',
    'DACT_dedo1_pie_izquierdo', 'DACT_dedo2_pie_izquierdo', 'DACT_dedo3_pie_izquierdo', 'DACT_dedo4_pie_izquierdo', 'DACT_dedo5_pie_izquierdo',
    # Totales
    'NAD_Total', 'NAT_Total', 'Dactilitis_Total',
    # Antropometría
    'Peso', 'Talla', 'IMC', 'TA',
    # PROs
    'EVA_Global', 'EVA_Dolor', 'EVA_Fatiga', 'Rigidez_Matutina_Min', 'Dolor_Nocturno',
    # Afectación Psoriasis
    'Psoriasis_Cuero_Cabelludo', 'Psoriasis_Ungueal', 'Psoriasis_Extensora', 'Psoriasis_Pliegues', 'Psoriasis_Palmoplantar',
    # Manifestaciones Extraarticulares
    'ExtraArticular_Digestiva', 'ExtraArticular_Uveitis', 'ExtraArticular_Psoriasis',
    # Comorbilidades
    'Comorbilidad_HTA', 'Comorbilidad_DM', 'Comorbilidad_DLP', 'Comorbilidad_ECV', 'Comorbilidad_Gastritis', 'Comorbilidad_Obesidad', 'Comorbilidad_Osteoporosis', 'Comorbilidad_Gota',
    # Antecedentes Familiares
    'AF_Psoriasis', 'AF_Artritis', 'AF_EII', 'AF_Uveitis',
    # Tóxicos
    'Toxico_Tabaco', 'Toxico_Tabaco_Desc', 'Toxico_Alcohol', 'Toxico_Alcohol_Desc', 'Toxico_Drogas', 'Toxico_Drogas_Desc',
    # Entesitis
    'Entesitis_Aquiles_Der', 'Entesitis_Fascia_Der', 'Entesitis_Epicondilo_Lat_Der', 'Entesitis_Epicondilo_Med_Der', 'Entesitis_Trocanter_Der',
    'Entesitis_Aquiles_Izq', 'Entesitis_Fascia_Izq', 'Entesitis_Epicondilo_Lat_Izq', 'Entesitis_Epicondilo_Med_Izq', 'Entesitis_Trocanter_Izq', 'Otras_Entesitis',
    # Pruebas complementarias
    'PCR', 'VSG', 'Otros_Hallazgos_Analitica', 'Hallazgos_Radiografia', 'Hallazgos_RMN',
    # BASDAI
    'BASDAI_P1', 'BASDAI_P2', 'BASDAI_P3', 'BASDAI_P4', 'BASDAI_P5', 'BASDAI_P6', 'BASDAI_Result',
    # ASDAS
    'ASDAS_Dolor_Espalda', 'ASDAS_Duracion_Rigidez', 'ASDAS_EVA_Global', 'ASDAS_CRP_Result', 'ASDAS_ESR_Result',
    # Metrología
    'Schober', 'Rotacion_Cervical', 'Distancia_OP', 'Distancia_TP', 'Expansion_Toracica', 'Distancia_Intermaleolar',
    # Evaluación Psoriasis
    'PASI_Score', 'BSA_Percentage', 'Psoriasis_Descripcion',
    # HAQ-DI
    'HAQ_Vestirse', 'HAQ_Levantarse', 'HAQ_Comer', 'HAQ_Caminar', 'HAQ_Higiene', 'HAQ_Alcanzar', 'HAQ_Agarrar', 'HAQ_Actividades', 'HAQ_Total',
    # LEI
    'LEI_Epicondilo_Lat_Izq', 'LEI_Epicondilo_Lat_Der', 'LEI_Epicondilo_Med_Izq', 'LEI_Epicondilo_Med_Der', 'LEI_Aquiles_Izq', 'LEI_Aquiles_Der', 'LEI_Score',
    # MDA
    'MDA_NAT', 'MDA_NAD', 'MDA_PASI', 'MDA_Dolor', 'MDA_Global', 'MDA_HAQ', 'MDA_Entesitis', 'MDA_Cumple',
    # RAPID3
    'RAPID3_Funcion', 'RAPID3_Dolor', 'RAPID3_Global', 'RAPID3_Score',
    # Tratamiento Actual
    'Tratamiento_Actual', 'Fecha_Inicio_Tratamiento', 'Decision_Terapeutica_PV',
    # Continuar Tratamiento
    'Continuar_Adherencia', 'Continuar_Ajuste_Terapeutico',
    # Cambio Tratamiento
    'Cambio_Motivo', 'Cambio_Efectos_Adversos', 'Cambio_Descripcion_Efectos',
    'Cambio_Sistemico_Farmaco', 'Cambio_Sistemico_Dosis', 'Cambio_FAME_Farmaco', 'Cambio_FAME_Dosis', 'Cambio_Biologico_Farmaco', 'Cambio_Biologico_Dosis',
    # Decision Terapeutica Seguimiento
    'Decision_Terapeutica_SEG',
    # Tratamientos Iniciales
    'Trat_Sistemico', 'Trat_Sistemico_Dosis', 'Trat_FAME', 'Trat_FAME_Dosis', 'Trat_Biologico', 'Trat_Biologico_Dosis',
    # Seguimiento
    'Fecha_Proxima_Revision', 'Comentarios_Adicionales'
]

print(f"Total de columnas: {len(headers)}")

# Crear hoja ESPA
ws_espa = wb.create_sheet('ESPA', 0)
# Crear hoja APS
ws_aps = wb.create_sheet('APS', 1)

# Estilos para headers
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Añadir headers a ESPA
for col_num, header in enumerate(headers, 1):
    cell = ws_espa.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    ws_espa.column_dimensions[get_column_letter(col_num)].width = 13

# Añadir headers a APS
for col_num, header in enumerate(headers, 1):
    cell = ws_aps.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    ws_aps.column_dimensions[get_column_letter(col_num)].width = 13

# Crear hoja Fármacos
ws_farmacos = wb.create_sheet('Fármacos', 2)
ws_farmacos.cell(row=1, column=1).value = 'Sistemicos'
ws_farmacos.cell(row=1, column=2).value = 'FAMEs'
ws_farmacos.cell(row=1, column=3).value = 'Biologicos'
for col in [1, 2, 3]:
    cell = ws_farmacos.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    ws_farmacos.column_dimensions[get_column_letter(col)].width = 25

# Crear hoja Profesionales
ws_profesionales = wb.create_sheet('Profesionales', 3)
ws_profesionales.cell(row=1, column=1).value = 'Nombre'
ws_profesionales.cell(row=1, column=2).value = 'Cargo'
for col in [1, 2]:
    cell = ws_profesionales.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    ws_profesionales.column_dimensions[get_column_letter(col)].width = 25

# Guardar el archivo
wb.save('Hub_Clinico_Maestro.xlsx')
print("Exito: Hub_Clinico_Maestro.xlsx creado")
